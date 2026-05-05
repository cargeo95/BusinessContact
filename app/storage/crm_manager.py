from __future__ import annotations

import os
import shutil
from dataclasses import dataclass
from pathlib import Path

from app.utils.domain_utils import normalize_domain
from app.utils.url_utils import canonicalize_linkedin_company_url, normalize_url

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - dependency is added later in the project sequence
    Workbook = None
    load_workbook = None

_MAX_BACKUPS = 3


CRM_HEADERS = [
    "empresa",
    "linkedin_url",
    "dominio",
    "correo_1",
    "correo_2",
    "correo_3",
    "telefono_1",
    "telefono_2",
    "pais",
    "fecha",
    "estado",
    "detalle",
]


@dataclass(frozen=True)
class CRMRecord:
    empresa: str
    linkedin_url: str = ""
    dominio: str = ""
    correo_1: str = ""
    correo_2: str = ""
    correo_3: str = ""
    telefono_1: str = ""
    telefono_2: str = ""
    pais: str = ""
    fecha: str = ""
    estado: str = ""
    detalle: str = ""

    def as_row(self) -> list[str]:
        return [
            self.empresa,
            self.linkedin_url,
            self.dominio,
            self.correo_1,
            self.correo_2,
            self.correo_3,
            self.telefono_1,
            self.telefono_2,
            self.pais,
            self.fecha,
            self.estado,
            self.detalle,
        ]


class CRMManager:
    def __init__(self, crm_path: Path) -> None:
        self.crm_path = crm_path
        self.crm_path.parent.mkdir(parents=True, exist_ok=True)
        self._cached_domains: set[str] | None = None
        self._cached_linkedin_urls: set[str] | None = None

    def ensure_workbook_exists(self) -> None:
        self._require_openpyxl()
        if self.crm_path.exists():
            self._migrate_workbook_if_needed()
            return

        workbook = Workbook()  # type: ignore[misc]
        try:
            sheet = workbook.active
            sheet.title = "CRM"
            self._write_headers(sheet)
            self._safe_save(workbook)
        finally:
            if hasattr(workbook, "close"):
                workbook.close()

    def load_existing_domains(self) -> set[str]:
        if self._cached_domains is not None:
            return set(self._cached_domains)

        if not self.crm_path.exists():
            self._cached_domains = set()
            return set()

        self._require_openpyxl()
        workbook = load_workbook(self.crm_path, read_only=True)  # type: ignore[misc]
        try:
            sheet = workbook.active
            header_map = self._build_header_map(sheet[1])
            domain_column = header_map.get("dominio")
            status_column = header_map.get("estado")

            domains: set[str] = set()
            if domain_column is None:
                self._cached_domains = domains
                return set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                status = (
                    str(row[status_column] or "").strip().lower()
                    if status_column is not None and row and len(row) > status_column
                    else ""
                )
                if status and status != "saved":
                    continue

                raw_domain = row[domain_column] if row and len(row) > domain_column else None
                domain = normalize_domain(str(raw_domain or ""))
                if domain:
                    domains.add(domain)

            self._cached_domains = domains
            return set(domains)
        finally:
            if hasattr(workbook, "close"):
                workbook.close()

    def domain_exists(self, domain: str) -> bool:
        return normalize_domain(domain) in self.load_existing_domains()

    def load_attempted_linkedin_urls(self) -> set[str]:
        if self._cached_linkedin_urls is not None:
            return set(self._cached_linkedin_urls)

        if not self.crm_path.exists():
            self._cached_linkedin_urls = set()
            return set()

        self._require_openpyxl()
        workbook = load_workbook(self.crm_path, read_only=True)  # type: ignore[misc]
        try:
            sheet = workbook.active
            header_map = self._build_header_map(sheet[1])
            linkedin_column = header_map.get("linkedin_url")
            urls: set[str] = set()
            if linkedin_column is None:
                self._cached_linkedin_urls = urls
                return set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                raw_value = row[linkedin_column] if row and len(row) > linkedin_column else None
                linkedin_url = canonicalize_linkedin_company_url(str(raw_value or "")) or normalize_url(
                    str(raw_value or "")
                )
                if linkedin_url:
                    urls.add(linkedin_url)

            self._cached_linkedin_urls = urls
            return set(urls)
        finally:
            if hasattr(workbook, "close"):
                workbook.close()

    def attempt_exists(self, linkedin_url: str) -> bool:
        normalized_url = canonicalize_linkedin_company_url(linkedin_url) or normalize_url(linkedin_url)
        if not normalized_url:
            return False
        return normalized_url in self.load_attempted_linkedin_urls()

    def upsert_record(self, record: CRMRecord) -> None:
        self.ensure_workbook_exists()
        self._require_openpyxl()
        workbook, sheet = self._open_workbook()

        try:
            normalized_record = CRMRecord(
                empresa=record.empresa,
                linkedin_url=canonicalize_linkedin_company_url(record.linkedin_url)
                or normalize_url(record.linkedin_url),
                dominio=normalize_domain(record.dominio),
                correo_1=record.correo_1,
                correo_2=record.correo_2,
                correo_3=record.correo_3,
                pais=record.pais,
                fecha=record.fecha,
                estado=record.estado.strip().lower(),
                detalle=record.detalle,
            )

            existing_row_index = self._find_row_by_linkedin_url(sheet, normalized_record.linkedin_url)
            if existing_row_index is None:
                sheet.append(normalized_record.as_row())
            else:
                for column_index, value in enumerate(normalized_record.as_row(), start=1):
                    sheet.cell(row=existing_row_index, column=column_index, value=value)

            self._safe_save(workbook)
            self._cached_domains = None
            self._cached_linkedin_urls = None
        finally:
            if hasattr(workbook, "close"):
                workbook.close()

    def load_rows(self) -> list[CRMRecord]:
        if not self.crm_path.exists():
            return []

        self._require_openpyxl()
        workbook = load_workbook(self.crm_path, read_only=True)  # type: ignore[misc]
        try:
            sheet = workbook.active
            rows: list[CRMRecord] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                padded = list(values or ()) + [""] * max(0, len(CRM_HEADERS) - len(values or ()))
                rows.append(
                    CRMRecord(
                        empresa=str(padded[0] or ""),
                        linkedin_url=str(padded[1] or ""),
                        dominio=str(padded[2] or ""),
                        correo_1=str(padded[3] or ""),
                        correo_2=str(padded[4] or ""),
                        correo_3=str(padded[5] or ""),
                        telefono_1=str(padded[6] or ""),
                        telefono_2=str(padded[7] or ""),
                        pais=str(padded[8] or ""),
                        fecha=str(padded[9] or ""),
                        estado=str(padded[10] or ""),
                        detalle=str(padded[11] or ""),
                    )
                )
            return rows
        finally:
            if hasattr(workbook, "close"):
                workbook.close()

    def reset_workbook(self) -> None:
        self._require_openpyxl()
        workbook = Workbook()  # type: ignore[misc]
        try:
            sheet = workbook.active
            sheet.title = "CRM"
            self._write_headers(sheet)
            self._safe_save(workbook)
            self._cached_domains = set()
            self._cached_linkedin_urls = set()
        finally:
            if hasattr(workbook, "close"):
                workbook.close()

    def _open_workbook(self):
        if self.crm_path.exists():
            workbook = load_workbook(self.crm_path)  # type: ignore[misc]
            sheet = workbook.active
        else:
            workbook = Workbook()  # type: ignore[misc]
            sheet = workbook.active
            sheet.title = "CRM"
            self._write_headers(sheet)
        return workbook, sheet

    @staticmethod
    def _build_header_map(header_row) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for index, cell in enumerate(header_row):
            header_value = str(getattr(cell, "value", "") or "").strip().lower()
            if header_value:
                header_map[header_value] = index
        return header_map

    @staticmethod
    def _write_headers(sheet) -> None:
        for index, header in enumerate(CRM_HEADERS, start=1):
            sheet.cell(row=1, column=index, value=header)

    @staticmethod
    def _find_row_by_linkedin_url(sheet, linkedin_url: str) -> int | None:
        if not linkedin_url:
            return None

        for row_index in range(2, sheet.max_row + 1):
            current_value = normalize_url(str(sheet.cell(row=row_index, column=2).value or ""))
            if not current_value:
                continue
            current_value = canonicalize_linkedin_company_url(current_value) or current_value
            if current_value == linkedin_url:
                return row_index
        return None

    def _migrate_workbook_if_needed(self) -> None:
        workbook = load_workbook(self.crm_path)  # type: ignore[misc]
        try:
            sheet = workbook.active
            if sheet.max_row == 0:
                self._write_headers(sheet)
                self._safe_save(workbook)
                return

            existing_headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
            if existing_headers == CRM_HEADERS:
                return

            existing_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        finally:
            if hasattr(workbook, "close"):
                workbook.close()

        normalized_rows = self._normalize_rows_for_migration(existing_headers, existing_rows)
        migrated_workbook = Workbook()  # type: ignore[misc]
        try:
            migrated_sheet = migrated_workbook.active
            migrated_sheet.title = "CRM"
            self._write_headers(migrated_sheet)
            for row in normalized_rows:
                migrated_sheet.append(row.as_row())
            self._safe_save(migrated_workbook)
            self._cached_domains = None
            self._cached_linkedin_urls = None
        finally:
            if hasattr(migrated_workbook, "close"):
                migrated_workbook.close()

    @staticmethod
    def _normalize_rows_for_migration(headers: list[str], rows: list[tuple]) -> list[CRMRecord]:
        header_map = {str(header or "").strip().lower(): index for index, header in enumerate(headers)}
        normalized_records: list[CRMRecord] = []

        for row in rows:
            values = list(row or ())

            def get_value(column_name: str) -> str:
                index = header_map.get(column_name)
                if index is None or index >= len(values):
                    return ""
                return str(values[index] or "")

            correo_1 = get_value("correo_1")
            correo_2 = get_value("correo_2")
            correo_3 = get_value("correo_3")
            estado = get_value("estado").strip().lower()
            if not estado:
                estado = "saved" if any([correo_1, correo_2, correo_3]) else ""

            normalized_records.append(
                CRMRecord(
                    empresa=get_value("empresa"),
                    linkedin_url=canonicalize_linkedin_company_url(get_value("linkedin_url"))
                    or normalize_url(get_value("linkedin_url")),
                    dominio=normalize_domain(get_value("dominio")),
                    correo_1=correo_1,
                    correo_2=correo_2,
                    correo_3=correo_3,
                    telefono_1=get_value("telefono_1"),
                    telefono_2=get_value("telefono_2"),
                    pais=get_value("pais"),
                    fecha=get_value("fecha"),
                    estado=estado,
                    detalle=get_value("detalle"),
                )
            )

        return normalized_records

    def _safe_save(self, workbook) -> None:
        """Save via temp file + atomic rename, then rotate backups."""
        tmp = self.crm_path.with_suffix(".tmp")
        workbook.save(tmp)

        # Rotate backups: crm.xlsx.bak1 → .bak2 → .bak3 (oldest dropped)
        if self.crm_path.exists():
            for i in range(_MAX_BACKUPS - 1, 0, -1):
                older = self.crm_path.with_suffix(f".bak{i + 1}")
                newer = self.crm_path.with_suffix(f".bak{i}")
                if newer.exists():
                    shutil.copy2(newer, older)
            shutil.copy2(self.crm_path, self.crm_path.with_suffix(".bak1"))

        os.replace(tmp, self.crm_path)

    def update_phones_by_domain(self, domain: str, telefono_1: str, telefono_2: str) -> bool:
        """Update telefono_1 and telefono_2 for the row matching domain. Returns True if found."""
        self.ensure_workbook_exists()
        self._require_openpyxl()
        workbook, sheet = self._open_workbook()

        try:
            header_map = self._build_header_map(sheet[1])
            domain_col = header_map.get("dominio")
            t1_col = header_map.get("telefono_1")
            t2_col = header_map.get("telefono_2")

            if domain_col is None:
                return False

            normalized = normalize_domain(domain)
            for row_idx in range(2, sheet.max_row + 1):
                cell_val = str(sheet.cell(row=row_idx, column=domain_col + 1).value or "").strip().lower()
                if cell_val == normalized:
                    if t1_col is not None:
                        sheet.cell(row=row_idx, column=t1_col + 1, value=telefono_1)
                    if t2_col is not None:
                        sheet.cell(row=row_idx, column=t2_col + 1, value=telefono_2)
                    self._safe_save(workbook)
                    return True
            return False
        finally:
            if hasattr(workbook, "close"):
                workbook.close()

    @staticmethod
    def _require_openpyxl() -> None:
        if Workbook is None or load_workbook is None:
            raise RuntimeError(
                "CRMManager necesita openpyxl para manejar crm.xlsx. "
                "La dependencia se agregara cuando lleguemos al paso 24."
            )
