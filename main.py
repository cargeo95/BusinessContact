from __future__ import annotations

import csv
import sys
import time
from datetime import datetime
from pathlib import Path

from app.core.bootstrap import bootstrap_application
from app.core.agent_runner import AgentRunSummary
from app.core.config import AppConfig

# Forzar UTF-8 en stdout/stderr para no fallar con caracteres especiales en Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def _update_plan(config: AppConfig, summary: AgentRunSummary) -> tuple[int, str]:
    """Marca el query actual como usado en queries_plan.xlsx. Devuelve (numero, siguiente_query)."""
    plan_path = config.project_root / "data" / "queries_plan.xlsx"
    if not plan_path.exists():
        return 0, ""

    try:
        from openpyxl import load_workbook
        wb = load_workbook(plan_path)
        ws = wb.active

        current = config.maps_search_query.strip().lower()
        query_number = 0
        next_query = ""
        found = False

        for row in ws.iter_rows(min_row=2):
            cell_query = str(row[1].value or "").strip()   # col B
            used = str(row[4].value or "").strip().lower()  # col E

            if cell_query.lower() == current:
                found = True
                query_number = int(row[0].value or 0)
                row[4].value = "Si"                                          # col E Usado
                row[5].value = datetime.now().strftime("%Y-%m-%d %H:%M")    # col F Fecha
                row[6].value = summary.companies_seen                        # col G Vistas
                row[7].value = summary.companies_processed                   # col H Procesadas
            elif found and used not in ("si", "sí") and not next_query:
                next_query = cell_query

        wb.save(plan_path)
        return query_number, next_query
    except Exception:
        return 0, ""


def _log_csv(config: AppConfig, summary: AgentRunSummary) -> None:
    log_path = config.project_root / "data" / "queries_log.csv"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    is_new = not log_path.exists()
    with open(log_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if is_new:
            writer.writerow(["fecha", "query", "vistas", "procesadas", "omitidas"])
        writer.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            config.maps_search_query,
            summary.companies_seen,
            summary.companies_processed,
            summary.companies_skipped,
        ])


def _advance_env_query(config: AppConfig, next_query: str) -> None:
    """Reemplaza MAPS_SEARCH_QUERY en el .env con el siguiente query del plan."""
    env_path = config.project_root / ".env"
    if not env_path.exists() or not next_query:
        return

    lines = env_path.read_text(encoding="utf-8").splitlines()
    updated = []
    replaced = False
    for line in lines:
        if line.strip().upper().startswith("MAPS_SEARCH_QUERY"):
            updated.append(f"MAPS_SEARCH_QUERY={next_query}")
            replaced = True
        else:
            updated.append(line)

    if not replaced:
        updated.append(f"MAPS_SEARCH_QUERY={next_query}")

    env_path.write_text("\n".join(updated) + "\n", encoding="utf-8")


def _print_summary(config: AppConfig, summary: AgentRunSummary, query_number: int, next_query: str) -> None:
    sep = "=" * 62
    line = "-" * 62
    label = f"QUERY #{query_number}" if query_number else "RESULTADO"
    print(f"\n{sep}")
    print(f"  {label}  —  {config.maps_search_query}")
    print(line)
    print(f"  Vistas:      {summary.companies_seen}")
    print(f"  Procesadas:  {summary.companies_processed}  (con correos guardados)")
    print(f"  Omitidas:    {summary.companies_skipped}")
    if summary.messages:
        print(line)
        for msg in summary.messages:
            print(f"  · {msg}")
    print(line)
    if next_query:
        print(f"  .env actualizado automaticamente con query #{query_number + 1}:")
        print(f"  {next_query}")
    else:
        print("  No hay mas queries pendientes en el plan.")
    print(f"{sep}\n")


def main() -> int:
    while True:
        application = None
        try:
            application = bootstrap_application()
            summary = application.runner.run()
            application.finalizer.finalize(summary=summary)
            query_number, next_query = _update_plan(application.config, summary)
            _log_csv(application.config, summary)
            _advance_env_query(application.config, next_query)
            _print_summary(application.config, summary, query_number, next_query)

            if not next_query:
                print("Plan completado. Todos los queries ejecutados.")
                return 0

            time.sleep(3)

        except Exception as exc:
            if application is not None:
                application.finalizer.finalize(error=exc)
            print(f"\nEjecucion fallida: {exc}")
            return 1


if __name__ == "__main__":
    raise SystemExit(main())
